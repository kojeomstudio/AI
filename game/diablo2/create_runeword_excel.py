import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "D2R 룬워드 목록"

# Styles
header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
data_font = Font(name="맑은 고딕", size=10)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Category fills
cat_109 = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # green
cat_110 = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # blue
cat_111 = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # orange
cat_ladder = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # yellow
cat_24 = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")  # purple
cat_26 = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")  # red

# Headers
headers = ["No.", "한글 이름", "영문 이름 (Runeword)", "룬 조합 순서", "필요 아이템", "소켓 수", "추가 버전"]
col_widths = [6, 18, 26, 36, 30, 10, 14]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Runeword data: (한글명, 영문명, 룬 조합, 필요 아이템, 소켓수, 버전)
runewords = [
    # === 1.09 ===
    ("강철", "Steel", "Tir + El", "도검 / 도끼 / 철퇴", 2, "1.09"),
    ("구렁텅이", "Nadir", "Nef + Tir", "투구", 2, "1.09"),
    ("잎새", "Leaf", "Tir + Ral", "양손 지팡이 (소서리스 전용)", 2, "1.09"),
    ("잠행", "Stealth", "Tal + Eth", "갑옷", 2, "1.09"),
    ("강함", "Strength", "Amn + Tir", "근접 무기", 2, "1.09"),
    ("전승", "Lore", "Ort + Sol", "투구", 2, "1.09"),
    ("각운", "Rhyme", "Shael + Eth", "방패", 2, "1.09"),
    ("서풍", "Zephyr", "Ort + Eth", "원거리 무기 (활/석궁)", 2, "1.09"),
    ("순백", "White", "Dol + Io", "완드 (네크로맨서 전용)", 2, "1.09"),
    ("연기", "Smoke", "Nef + Lum", "갑옷", 2, "1.09"),
    ("악의", "Malice", "Ith + El + Eth", "근접 무기", 3, "1.09"),
    ("고대인의 서약", "Ancient's Pledge", "Ral + Ort + Tal", "방패", 3, "1.09"),
    ("왕의 은총", "King's Grace", "Amn + Ral + Thul", "도검 / 홀", 3, "1.09"),
    ("광휘", "Radiance", "Nef + Sol + Ith", "투구", 3, "1.09"),
    ("어둠", "Black", "Thul + Io + Nef", "곤봉 / 철퇴 / 망치", 3, "1.09"),
    ("선율", "Melody", "Shael + Ko + Nef", "원거리 무기 (활/석궁)", 3, "1.09"),
    ("용맹", "Lionheart", "Hel + Lum + Fal", "갑옷", 3, "1.09"),
    ("부", "Wealth", "Lem + Ko + Tir", "갑옷", 3, "1.09"),
    ("맹독", "Venom", "Tal + Dol + Mal", "무기", 3, "1.09"),
    ("분노", "Fury", "Jah + Gul + Eth", "근접 무기", 3, "1.09"),
    ("신성한 천둥", "Holy Thunder", "Eth + Ral + Ort + Tal", "홀 (성기사 전용)", 4, "1.09"),
    ("기억", "Memory", "Lum + Io + Sol + Eth", "양손 지팡이 (소서리스 전용)", 4, "1.09"),
    ("명예", "Honor", "Amn + El + Ith + Tir + Sol", "근접 무기", 5, "1.09"),
    ("침묵", "Silence", "Dol + Eld + Hel + Ist + Tir + Vex", "무기", 6, "1.09"),

    # === 1.10 ===
    ("광채", "Splendor", "Eth + Lum", "방패", 2, "1.10"),
    ("신중", "Prudence", "Mal + Tir", "갑옷", 2, "1.10"),
    ("바람", "Wind", "Sur + El", "근접 무기", 2, "1.10"),
    ("수수께끼", "Enigma", "Jah + Ith + Ber", "갑옷", 3, "1.10"),
    ("협박", "Duress", "Shael + Um + Thul", "갑옷", 3, "1.10"),
    ("어스름", "Gloom", "Fal + Um + Pul", "갑옷", 3, "1.10"),
    ("초승달", "Crescent Moon", "Shael + Um + Tir", "도검 / 도끼 / 미늘창", 3, "1.10"),
    ("성역", "Sanctuary", "Ko + Ko + Mal", "방패", 3, "1.10"),
    ("착란", "Delirium", "Lem + Ist + Io", "투구", 3, "1.10"),
    ("혼돈", "Chaos", "Fal + Ohm + Um", "손톱 (어쌔신 전용)", 3, "1.10"),
    ("열정", "Passion", "Dol + Ort + Eld + Lem", "무기", 4, "1.10"),
    ("돌", "Stone", "Shael + Um + Pul + Lum", "갑옷", 4, "1.10"),
    ("왕 시해자", "Kingslayer", "Mal + Um + Gul + Fal", "도검 / 도끼", 4, "1.10"),
    ("참나무의 심장", "Heart of the Oak", "Ko + Vex + Pul + Thul", "철퇴 / 지팡이", 4, "1.10"),
    ("추방", "Exile", "Vex + Ohm + Ist + Dol", "성기사 전용 방패", 4, "1.10"),
    ("찔레", "Bramble", "Ral + Ohm + Sur + Eth", "갑옷", 4, "1.10"),
    ("명예의 굴레", "Chains of Honor", "Dol + Um + Ber + Ist", "갑옷", 4, "1.10"),
    ("기근", "Famine", "Fal + Ohm + Ort + Jah", "도끼 / 망치", 4, "1.10"),
    ("정의의 손길", "Hand of Justice", "Sur + Cham + Amn + Lo", "무기", 4, "1.10"),
    ("소집", "Call to Arms", "Amn + Ral + Mal + Ist + Ohm", "무기", 5, "1.10"),
    ("야수", "Beast", "Ber + Tir + Um + Mal + Lum", "도끼 / 망치 / 홀", 5, "1.10"),
    ("영원", "Eternity", "Amn + Ber + Ist + Sol + Sur", "근접 무기", 5, "1.10"),
    ("파멸", "Doom", "Hel + Ohm + Um + Lo + Cham", "도끼 / 망치 / 미늘창", 5, "1.10"),
    ("죽어가는 자의 숨결", "Breath of the Dying", "Vex + Hel + El + Eld + Zod + Eth", "무기", 6, "1.10"),

    # === 1.11 ===
    ("신화", "Myth", "Hel + Amn + Nef", "갑옷 (바바리안 전용)", 3, "1.11"),
    ("평화", "Peace", "Shael + Thul + Amn", "갑옷 (아마존 전용)", 3, "1.11"),
    ("배신", "Treachery", "Shael + Thul + Lem", "갑옷", 3, "1.11"),
    ("깨우침", "Enlightenment", "Pul + Ral + Sol", "갑옷 (소서리스 전용)", 3, "1.11"),
    ("뼈", "Bone", "Sol + Um + Um", "갑옷 (네크로맨서 전용)", 3, "1.11"),
    ("비", "Rain", "Ort + Mal + Ith", "갑옷 (드루이드 전용)", 3, "1.11"),
    ("원칙", "Principle", "Ral + Gul + Eld", "갑옷 (성기사 전용)", 3, "1.11"),

    # === 구 래더 전용 (현재 모두 해제) ===
    ("모서리", "Edge", "Tir + Tal + Amn", "원거리 무기 (활/석궁)", 3, "래더→해제"),
    ("집행자", "Lawbringer", "Amn + Lem + Ko", "도검 / 망치 / 홀", 3, "래더→해제"),
    ("용", "Dragon", "Sur + Lo + Sol", "갑옷 / 방패", 3, "래더→해제"),
    ("꿈", "Dream", "Io + Jah + Pul", "투구 / 방패", 3, "래더→해제"),
    ("정령", "Spirit", "Tal + Thul + Ort + Amn", "도검 / 방패", 4, "래더→해제"),
    ("통찰", "Insight", "Ral + Tir + Tal + Sol", "미늘창 / 지팡이", 4, "래더→해제"),
    ("조화", "Harmony", "Tir + Ith + Sol + Ko", "원거리 무기 (활/석궁)", 4, "래더→해제"),
    ("불굴", "Fortitude", "El + Sol + Dol + Lo", "무기 / 갑옷", 4, "래더→해제"),
    ("무한", "Infinity", "Ber + Mal + Ber + Ist", "미늘창", 4, "래더→해제"),
    ("불사조", "Phoenix", "Vex + Vex + Lo + Jah", "무기 / 방패", 4, "래더→해제"),
    ("얼음", "Ice", "Amn + Shael + Jah + Lo", "원거리 무기 (활/석궁)", 4, "래더→해제"),
    ("맹세", "Oath", "Shael + Pul + Mal + Lum", "도끼 / 철퇴 / 도검", 4, "래더→해제"),
    ("균열", "Rift", "Hel + Ko + Lem + Gul", "미늘창 / 홀", 4, "래더→해제"),
    ("각인", "Brand", "Jah + Lo + Mal + Gul", "원거리 무기 (활/석궁)", 4, "래더→해제"),
    ("믿음", "Faith", "Ohm + Jah + Lem + Eld", "원거리 무기 (활/석궁)", 4, "래더→해제"),
    ("긍지", "Pride", "Cham + Sur + Io + Lo", "미늘창", 4, "래더→해제"),
    ("분별", "Voice of Reason", "Lem + Ko + El + Eld", "철퇴 / 도검", 4, "래더→해제"),
    ("복종", "Obedience", "Hel + Ko + Thul + Eth + Fal", "미늘창", 5, "래더→해제"),
    ("죽음", "Death", "Hel + El + Vex + Ort + Gul", "도검 / 도끼", 5, "래더→해제"),
    ("슬픔", "Grief", "Eth + Tir + Lo + Mal + Ral", "도검 / 도끼", 5, "래더→해제"),
    ("파괴", "Destruction", "Vex + Lo + Ber + Jah + Ko", "미늘창 / 도검", 5, "래더→해제"),
    ("분노", "Wrath", "Pul + Lum + Ber + Mal", "원거리 무기 (활/석궁)", 4, "래더→해제"),
    ("마지막 소원", "Last Wish", "Jah + Mal + Jah + Sur + Jah + Ber", "도끼 / 망치 / 도검", 6, "래더→해제"),

    # === 2.4 패치 추가 ===
    ("꺼져가는 불꽃", "Flickering Flame", "Nef + Pul + Vex", "투구", 3, "2.4"),
    ("무늬", "Pattern", "Tal + Ort + Thul", "손톱 (어쌔신 전용)", 3, "2.4"),
    ("역병", "Plague", "Cham + Shael + Um", "도검 / 손톱 / 단검", 3, "2.4"),
    ("지혜", "Wisdom", "Pul + Ith + Eld", "투구", 3, "2.4"),
    ("안개", "Mist", "Cham + Shael + Gul + Thul + Ith", "원거리 무기 (활/석궁)", 5, "2.4"),
    ("불굴의 의지", "Unbending Will", "Fal + Io + Ith + Eld + El + Hel", "도검", 6, "2.4"),
    ("집착", "Obsession", "Zod + Ist + Lem + Lum + Io + Nef", "지팡이", 6, "2.4"),

    # === 2.6 패치 추가 ===
    ("방벽", "Bulwark", "Shael + Io + Sol", "투구", 3, "2.6"),
    ("치유", "Cure", "Shael + Io + Tal", "투구", 3, "2.6"),
    ("접지", "Ground", "Shael + Io + Ort", "투구", 3, "2.6"),
    ("난로", "Hearth", "Shael + Io + Thul", "투구", 3, "2.6"),
    ("담금질", "Temper", "Shael + Io + Ral", "투구", 3, "2.6"),
    ("탈태", "Metamorphosis", "Io + Cham + Fal", "투구", 3, "2.6"),
    ("투지", "Hustle", "Shael + Ko + Eld", "무기 / 갑옷", 3, "2.6"),
    ("모자이크", "Mosaic", "Mal + Gul + Amn", "손톱 (어쌔신 전용)", 3, "2.6"),
]

# Write data
version_fills = {
    "1.09": cat_109,
    "1.10": cat_110,
    "1.11": cat_111,
    "래더→해제": cat_ladder,
    "2.4": cat_24,
    "2.6": cat_26,
}

for row_idx, (kr_name, en_name, runes, item_type, sockets, version) in enumerate(runewords, 2):
    row_data = [row_idx - 1, kr_name, en_name, runes, item_type, sockets, version]
    fill = version_fills.get(version, PatternFill())
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.border = thin_border
        cell.fill = fill
        if col_idx in (1, 6):
            cell.alignment = center_align
        elif col_idx == 7:
            cell.alignment = center_align
        else:
            cell.alignment = left_align

# Freeze top row
ws.freeze_panes = "A2"

# Auto filter
ws.auto_filter.ref = f"A1:G{len(runewords) + 1}"

# Legend sheet
ws2 = wb.create_sheet("범례")
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 40

legends = [
    ("버전 색상 범례", "", None),
    ("1.09", "오리지널 룬워드 (클래식)", cat_109),
    ("1.10", "확장팩 (Lord of Destruction) 추가", cat_110),
    ("1.11", "1.11 패치 추가 (직업 전용 갑옷 등)", cat_111),
    ("래더→해제", "구 래더 전용 → 레저렉션에서 해제", cat_ladder),
    ("2.4", "레저렉션 2.4 패치 추가", cat_24),
    ("2.6", "레저렉션 2.6 패치 추가", cat_26),
]

for row_idx, (label, desc, fill) in enumerate(legends, 1):
    c1 = ws2.cell(row=row_idx, column=1, value=label)
    c2 = ws2.cell(row=row_idx, column=2, value=desc)
    if row_idx == 1:
        c1.font = Font(name="맑은 고딕", bold=True, size=12)
        c2.font = Font(name="맑은 고딕", bold=True, size=12)
    else:
        c1.font = data_font
        c2.font = data_font
        if fill:
            c1.fill = fill
            c2.fill = fill

ws2.cell(row=9, column=1, value="참고사항").font = Font(name="맑은 고딕", bold=True, size=11)
ws2.cell(row=10, column=1, value="• 룬워드는 반드시 일반(회색) 등급 아이템에만 적용 가능").font = data_font
ws2.cell(row=11, column=1, value="• 마법/희귀/세트/고유 아이템에는 적용 불가").font = data_font
ws2.cell(row=12, column=1, value="• 룬은 반드시 지정된 순서대로 삽입해야 함").font = data_font
ws2.cell(row=13, column=1, value="• 소켓 수가 정확히 일치해야 함 (초과 불가)").font = data_font
for r in range(10, 14):
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)

output_path = r"C:\workspaces\diablo2\D2R_룬워드_목록.xlsx"
wb.save(output_path)
print(f"파일 생성 완료: {output_path}")
print(f"총 룬워드 수: {len(runewords)}개")
